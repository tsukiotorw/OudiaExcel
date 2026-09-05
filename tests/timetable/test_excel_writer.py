from openpyxl import load_workbook

from src.models.railway import Railway
from src.timetable.excel_writer import ExcelWriter
from src.timetable.station_timetable_generator import (
    StationTimetableGenerator,
)

from src.timetable.display_config import TimetableDisplayConfig


def test_write_station_timetable(
    parsed_railway: Railway,
    tmp_path,
) -> None:
    """駅時刻表をExcelファイルへ出力できること。"""
    station = parsed_railway.stations[1]

    generator = StationTimetableGenerator(parsed_railway)
    timetable = generator.generate(station)

    output_path = tmp_path / "station_timetable.xlsx"

    writer = ExcelWriter()
    writer.write(timetable, output_path)

    assert output_path.exists()

    workbook = load_workbook(output_path)

    assert timetable.station_name in workbook.sheetnames

    worksheet = workbook[timetable.station_name]

    assert worksheet["A1"].value == "B 時刻表"
    assert worksheet["A3"].value == "下り"


def test_write_station_timetable_layout(
    parsed_railway: Railway,
    tmp_path,
) -> None:
    """駅時刻表が2行/時間の左右配置で出力されること。"""
    station = parsed_railway.stations[1]

    generator = StationTimetableGenerator(parsed_railway)
    timetable = generator.generate(station)

    output_path = tmp_path / "station_timetable.xlsx"

    writer = ExcelWriter()
    writer.write(timetable, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook[timetable.station_name]

    # 下り
    assert worksheet["A3"].value == "下り"

    # 下りは4時から開始
    assert worksheet["A4"].value == 4
    assert worksheet["A6"].value == 5
    assert worksheet["A8"].value == 6

    # 上りの開始位置を確認
    down_width = max(
        len(hour.entries)
        for hour in timetable.down
    ) + 1

    up_start_column = (
        1
        + down_width
        + writer._DIRECTION_GAP
    )

    up_start_cell = worksheet.cell(
        row=3,
        column=up_start_column,
    )

    assert up_start_cell.value == "上り"

    # 上りも4時から開始
    up_hour_cell = worksheet.cell(
        row=4,
        column=up_start_column,
    )

    assert up_hour_cell.value == 4


def test_write_station_timetable_entry(
    parsed_railway: Railway,
    tmp_path,
) -> None:
    """列車の行先・種別・分がExcelへ出力されること。"""
    station = parsed_railway.stations[1]

    generator = StationTimetableGenerator(parsed_railway)
    timetable = generator.generate(station)

    output_path = tmp_path / "station_timetable.xlsx"

    writer = ExcelWriter()
    writer.write(timetable, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook[timetable.station_name]

    # 4時台の下りデータを確認
    hour = timetable.down[0]

    assert hour.hour == 4
    assert hour.entries

    entry = hour.entries[0]

    # A列が時なので、最初の列車はB列
    assert worksheet["B4"].value == (
        entry.destination + entry.train_type.short_name
    )
    assert worksheet["B5"].value == entry.minute


def test_write_train_type_colors(
    parsed_railway: Railway,
    tmp_path,
) -> None:
    """列車種別に応じて文字色が設定されること。"""
    station = parsed_railway.stations[1]

    generator = StationTimetableGenerator(parsed_railway)
    timetable = generator.generate(station)

    # 列車種別の文字色とセルの色塗りを設定する
    display_config = TimetableDisplayConfig(
        train_type_colors={
            0: "008000",  # 普通：緑文字
            1: "0000FF",  # 快速：青文字
        },
        train_type_fills={
            0: "E2F0D9",  # 普通：薄緑
            1: "DDEBF7",  # 快速：薄青
        },
        header_fill="D9EAD3",
        hour_fill="F2F2F2",
    )

    # ↓↓↓ ここにデバッグ出力を追加 ↓↓↓
    entry = timetable.down[0].entries[0]
    print(f"entry.train_type.index = {entry.train_type.index}")
    print(f"display_config.train_type_colors = {display_config.train_type_colors}")
    print(f"get_train_type_color result = {display_config.get_train_type_color(entry.train_type.index)}")
    # ↑↑↑ ここまで ↑↑↑

    output_path = tmp_path / "station_timetable.xlsx"

    print(f"output_path = {output_path}")

    writer = ExcelWriter(
        display_config=display_config,
    )

    writer.write(
        timetable,
        output_path,
    )

    workbook = load_workbook(output_path)
    worksheet = workbook[timetable.station_name]

    # 4時台の最初の下り列車
    entry = timetable.down[0].entries[0]

    expected_color = display_config.get_train_type_color(
        entry.train_type.index
    )

    # ↓↓↓ ここにもデバッグ出力を追加 ↓↓↓
    print(f"expected_color = {expected_color}")
    print(f"B4 font color = {worksheet['B4'].font.color}")
    # ↑↑↑ ここまで ↑↑↑

    assert worksheet["B4"].font.color.type == "rgb"
    assert worksheet["B4"].font.color.rgb == (
        "FF" + expected_color
    )
