from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models.timetable import (
    StationTimetable,
    TimetableHour,
)

from src.timetable.display_config import TimetableDisplayConfig

class ExcelWriter:
    """StationTimetableをExcelファイルへ出力する。"""

    _DIRECTION_GAP = 2
    _HOUR_ROWS = 2

    def __init__(
        self,
        display_config: TimetableDisplayConfig | None = None,
    ) -> None:
        self.display_config = (
            display_config
            if display_config is not None
            else TimetableDisplayConfig()
        )

    def write(
        self,
        timetable: StationTimetable,
        output_path: Path,
    ) -> None:
        """駅時刻表をExcelファイルとして出力する。"""
        workbook = Workbook()
        worksheet = workbook.active

        if worksheet is None:
            raise RuntimeError("ワークシートを取得できません。")

        worksheet.title = timetable.station_name

        self._write_title(
            worksheet=worksheet,
            station_name=timetable.station_name,
        )

        down_width = self._max_entry_count(timetable.down) + 1
        up_width = self._max_entry_count(timetable.up) + 1

        down_start_column = 1

        up_start_column = (
            down_start_column
            + down_width
            + self._DIRECTION_GAP
        )

        start_row = 3

        down_end_row = self._write_direction(
            worksheet=worksheet,
            row=start_row,
            start_column=down_start_column,
            title="下り",
            hours=timetable.down,
            width=down_width,
        )

        up_end_row = self._write_direction(
            worksheet=worksheet,
            row=start_row,
            start_column=up_start_column,
            title="上り",
            hours=timetable.up,
            width=up_width,
        )

        self._apply_layout(
            worksheet=worksheet,
            down_start_column=down_start_column,
            down_width=down_width,
            up_start_column=up_start_column,
            up_width=up_width,
            table_end_row=max(down_end_row, up_end_row),
        )

        workbook.save(output_path)

    @staticmethod
    def _max_entry_count(
        hours: list[TimetableHour],
    ) -> int:
        """1時間あたりの最大列車本数を求める。"""
        if not hours:
            return 0

        return max(
            len(hour.entries)
            for hour in hours
        )

    @staticmethod
    def _write_title(
        worksheet,
        station_name: str,
    ) -> None:
        """駅名タイトルを書き込む。"""
        cell = worksheet.cell(
            row=1,
            column=1,
            value=f"{station_name} 時刻表",
        )

        cell.font = Font(
            size=16,
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )


    def _write_direction(
        self,
        worksheet,
        row: int,
        start_column: int,
        title: str,
        hours: list[TimetableHour],
        width: int,
    ) -> int:
        """指定方向の駅時刻表を書き込む。"""
        end_column = start_column + width - 1

        # 方向見出し
        header_fill = self._create_fill(
            self.display_config.header_fill
        )

        for column in range(
            start_column,
            end_column + 1,
        ):
            cell = worksheet.cell(
                row=row,
                column=column,
            )

            cell.fill = header_fill or PatternFill()

            if column == start_column:
                cell.value = title
                cell.font = Font(
                    size=12,
                    bold=True,
                )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        worksheet.merge_cells(
            start_row=row,
            start_column=start_column,
            end_row=row,
            end_column=end_column,
        )

        row += 1

        for hour in hours:
            metadata_row = row
            minute_row = row + 1

            # 時
            hour_cell = worksheet.cell(
                row=metadata_row,
                column=start_column,
                value=hour.hour,
            )

            hour_cell.font = Font(
                size=14,
                bold=True,
            )

            hour_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            hour_fill = self._create_fill(
                self.display_config.hour_fill
            )

            if hour_fill is not None:
                worksheet.cell(
                    row=metadata_row,
                    column=start_column,
                ).fill = hour_fill

                worksheet.cell(
                    row=minute_row,
                    column=start_column,
                ).fill = hour_fill

            worksheet.merge_cells(
                start_row=metadata_row,
                start_column=start_column,
                end_row=minute_row,
                end_column=start_column,
            )

            # 行先・種別
            for column, entry in enumerate(
                hour.entries,
                start=start_column + 1,
            ):
                metadata = self._format_entry_metadata(entry)

                train_type_color = self._to_excel_color(
                    self.display_config.get_train_type_color(
                        entry.train_type.index
                    )
                )

                train_type_fill = self._create_fill(
                    self.display_config.get_train_type_fill(
                        entry.train_type.index
                    )
                )

                metadata_cell = worksheet.cell(
                    row=metadata_row,
                    column=column,
                    value=metadata,
                )

                metadata_cell.font = Font(
                    size=11,
                    color=train_type_color,
                )

                metadata_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

                # 分
                minute_cell = worksheet.cell(
                    row=minute_row,
                    column=column,
                    value=entry.minute,
                )

                minute_cell.font = Font(
                    size=14,
                    bold=True,
                )

                minute_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

                metadata_cell.fill = train_type_fill or PatternFill()
                minute_cell.fill = train_type_fill or PatternFill()

            row += self._HOUR_ROWS

        return row


    @staticmethod
    def _apply_layout(
        worksheet,
        down_start_column: int,
        down_width: int,
        up_start_column: int,
        up_width: int,
        table_end_row: int,
    ) -> None:
        """ワークシートの基本レイアウトを設定する。"""

        alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        thin = Side(style="thin")
        medium = Side(style="medium")

        for start_column, width in (
            (down_start_column, down_width),
            (up_start_column, up_width),
        ):
            end_column = start_column + width - 1

            # 時刻列
            worksheet.column_dimensions[
                get_column_letter(start_column)
            ].width = 4

            # 列車列
            for column in range(
                start_column + 1,
                end_column + 1,
            ):
                worksheet.column_dimensions[
                    get_column_letter(column)
                ].width = 7

            # 方向見出し
            worksheet.cell(
                row=3,
                column=start_column,
            ).border = Border(
                top=medium,
                bottom=medium,
                left=medium,
                right=medium,
            )

            # 方向見出しの結合セルにも罫線を設定
            for column in range(
                start_column,
                end_column + 1,
            ):
                worksheet.cell(
                    row=3,
                    column=column,
                ).border = Border(
                    top=medium,
                    bottom=medium,
                    left=medium if column == start_column else thin,
                    right=medium if column == end_column else thin,
                )

            # 時刻表本体
            for row in range(4, table_end_row):
                for column in range(
                    start_column,
                    end_column + 1,
                ):
                    cell = worksheet.cell(
                        row=row,
                        column=column,
                    )

                    cell.alignment = alignment

                    # 左右の縦罫線
                    left = (
                        medium
                        if column == start_column
                        else thin
                    )

                    right = (
                        medium
                        if column == end_column
                        else thin
                    )

                    # 2行1時間ごとの横罫線
                    hour_offset = row - 4

                    if hour_offset % 2 == 0:
                        top = medium
                    else:
                        top = thin

                    if hour_offset % 2 == 1:
                        bottom = thin
                    else:
                        bottom = thin

                    cell.border = Border(
                        left=left,
                        right=right,
                        top=top,
                        bottom=bottom,
                    )

            # 最終行の下端を太くする
            last_row = table_end_row - 1

            for column in range(
                start_column,
                end_column + 1,
            ):
                cell = worksheet.cell(
                    row=last_row,
                    column=column,
                )

                cell.border = Border(
                    left=(
                        medium
                        if column == start_column
                        else thin
                    ),
                    right=(
                        medium
                        if column == end_column
                        else thin
                    ),
                    top=cell.border.top,
                    bottom=medium,
                )

        # 行高
        worksheet.row_dimensions[1].height = 28
        worksheet.row_dimensions[3].height = 24

        for row in range(4, table_end_row):
            if (row - 4) % 2 == 0:
                # 行先・種別
                worksheet.row_dimensions[row].height = 18
            else:
                # 分
                worksheet.row_dimensions[row].height = 20

        # タイトル
        worksheet.row_dimensions[1].height = 28


    @staticmethod
    def _format_entry_metadata(entry) -> str:
        """列車の行先・種別を駅時刻表用の文字列に変換する。"""
        metadata = entry.destination

        if entry.train_type.short_name:
            metadata += f" {entry.train_type.short_name}"

        return metadata


    @staticmethod
    def _to_excel_color(
        color: str | None,
    ) -> str | None:
        """RGB色をExcel用ARGB色へ変換する。"""
        if color is None:
            return None

        if len(color) == 6:
            return f"FF{color}"

        if len(color) == 8:
            return color

        raise ValueError(
            f"色は6桁または8桁の16進数で指定してください: {color}"
        )


    @staticmethod
    def _create_fill(
        color: str | None,
    ) -> PatternFill | None:
        """塗りつぶし色からExcel用のFillを生成する。"""
        if color is None:
            return None

        return PatternFill(
            fill_type="solid",
            fgColor=ExcelWriter._to_excel_color(color),
        )

